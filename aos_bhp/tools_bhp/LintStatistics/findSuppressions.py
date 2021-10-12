#*******************************************************************************
#
# (C) COPYRIGHT Bombardier Transportation Sweden AB, 2019
#
# We reserve all rights in this file and in the information
# contained therein. Reproduction, use or disclosure to third
# parties without written authority is strictly forbidden.
#
#  DESCRIPTION:
#  Finds lint suppressions in source code and produces a summary.
#
#******************************************************************************

#******************************************************************************
#
# REVISION HISTORY :
#
# Date          Name        Changes
# ---------------------------------------------------------------------------
# 2019-06-03    csundin     Created
#
#******************************************************************************

#******************************************************************************
# INCLUDE FILES
#******************************************************************************
import os
import re
import csv
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment

def atoi(string):
    number = 0
    index = 0

    while index < len(string):
        char = ord(string[index])
        if char >= 0x30 and char <= 0x39:
            number = (char - 0x30) + number * 10
        else:
            break
        index += 1

    return number


def write_csv_file(errors_dict, output_filename):
    csv_file = open(output_filename,'w')

    wr = csv.writer(csv_file, dialect="excel", delimiter=",")

    wr.writerow(('Suppressed lint error', 'Occurs at', 'Motivation for suppression'))

    for suppression,occurences in sorted(errors_dict.items(), key=lambda t: atoi(t[0])):
        parts = suppression.split("|");
        error = parts[0]
        comment = parts[1]
        file_list = ""
        for occurence in sorted(occurences):
            file_list += occurence + " "

        complete_item = (error, file_list, comment)

        wr.writerow(complete_item)

    csv_file.close()


def write_excel_file(errors_dict, output_filename):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['C'].width = 70

    alignment=Alignment(horizontal='general',
                        vertical='top',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    ws.append(['Lint error', 'Occurs at', 'Reason for suppression'])
    line_no = 1

    for suppression,occurences in sorted(errors_dict.items(), key=lambda t: atoi(t[0])):
        parts = suppression.split("|");
        error = parts[0]
        comment = parts[1]
        file_list = ""
        for occurence in sorted(occurences):
            if file_list is not "":
                file_list += "\n"
            file_list += occurence

        ws.append([error, file_list, comment])
        line_no += 1

        ws['A'+str(line_no)].alignment = alignment;
        ws['B'+str(line_no)].alignment = alignment;
        ws['C'+str(line_no)].alignment = alignment;

    wb.save(output_filename)


def main():
    if len(sys.argv) != 3:
        print("Usage: ./findSuppressions.py <path_to_aos_bhp> <output_file>")
        print("where output_file must end with .csv or .xlsx")
        sys.exit(1)

    src_path = sys.argv[1]
    dst_file = sys.argv[2]

    sources = src_path + "/atc/impl " + src_path + "/atp_core/impl " + src_path + "/atp_bhp/impl"
    command = "find " + sources + " -type f -name \*.\[ch\]\* -exec grep -n lint /dev/null {} \;"

    print("Running: " + command)

    f = os.popen(command)

    errors_dict = defaultdict(set)

    print_output = False

    while True:
        line = f.readline()
        if not line: break

        match = re.search(r'(.*):(\d+):.*(/[/*]lint.*)', line)

        if match:
            src_file = match.group(1)
            src_line = match.group(2)
            code_line = match.group(3)

            if "atc/impl/simulation" in src_file or "atc/impl/vfw_sim" in src_file:
                continue

            parts = re.findall(r'( -e\d+| -estring)', code_line);
            if len(parts) > 0:
                print("error: global suppression found:")
                print("    " + line)

            parts = re.findall(r'( --e| -e| !e)([^ \(]+\([^"]+"[^"]+"\)|[^ \(]+\([^\)]+\)|{[^}]+}|\d+)', code_line);
            if len(parts) > 0:
                comment = code_line.replace("//lint", "")

                for part in parts:
                    suppression = "".join(part)
                    comment = comment.replace(suppression, "")

                for part in parts:
                    suppression = "".join(part)

                    match = re.search(r'(\d+).*', suppression)
                    if match:
                        error = match.group(1)

                        if error == "453":
                            error += ": Function 'X' previously designated pure"
                        elif error == "530":
                            error += ": Symbol 'X' not initialized"
                        elif error == "534":
                            error += ": Ignoring return value of function"
                        elif error == "551":
                            error += ": Symbol not accessed"
                        elif error == "571":
                            error += ": Suspicious cast"
                        elif error == "586":
                            error += ": Function 'X' is deprecated"
                        elif error == "701":
                            error += ": Shift left of signed quantity"
                        elif error == "714":
                            error += ": Symbol 'X' not referenced"
                        elif error == "716":
                            error += ": while(1) ..."
                        elif error == "717":
                            error += ": do ... while(0);"
                        elif error == "731":
                            error += ": Boolean argument to equal/not equal"
                        elif error == "734":
                            error += ": Loss of precision"
                        elif error == "751":
                            error += ": Local typedef 'X' not referenced"
                        elif error == "752":
                            error += ": Local declarator 'X' not referenced"
                        elif error == "759":
                            error += ": Declaration could be moved from header to module"
                        elif error == "765":
                            error += ": External 'X' could be made static"
                        elif error == "768":
                            error += ": Global struct member 'X' not referenced"
                        elif error == "769":
                            error += ": Global enumeration constant 'X' not referenced"
                        elif error == "818":
                            error += ": Pointer parameter could be declared as pointing to const"
                        elif error == "826":
                            error += ": Suspicious pointer-to-pointer conversion"
                        elif error == "829":
                            error += ": +headerwarn was previously issued for header 'X'"
                        elif error == "835":
                            error += ": A zero has been given as argument to operator"
                        elif error == "909":
                            error += ": Implicit conversion from int to bool"
                        elif error == "923":
                            error += ": Cast from int to pointer"
                        elif error == "925":
                            error += ": Cast from pointer to pointer"
                        elif error == "926":
                            error += ": Cast from pointer to pointer"
                        elif error == "927":
                            error += ": Cast from pointer to pointer"
                        elif error == "929":
                            error += ": Cast from pointer to pointer"
                        elif error == "946":
                            error += ": Relational or subtract operator applied to pointers"
                        elif error == "947":
                            error += ": Subtract operator applied to pointers"
                        elif error == "970":
                            error += ": Use of type 'X' outside of a typedef"
                        elif error == "1511":
                            error += ": Member hides non-virtual member in base class"
                        elif error == "1531":
                            error += ": Symbol 'X' should have compared argument against sizeof"
                        elif error == "1536":
                            error += ": Exposing low access member"
                        elif error == "1551":
                            error += ": Function may throw exception 'X' in destructor"
                        elif error == "1554":
                            error += ": Direct pointer copy of member"
                        elif error == "1555":
                            error += ": Direct pointer copy of member"
                        elif error == "1711":
                            error += ": Class has a virtual function but is not inherited"
                        elif error == "1714":
                            error += ": Member function 'X' not referenced"
                        elif error == "1716":
                            error += ": Virtual member function 'X' not referenced"
                        elif error == "1746":
                            error += ": Parameter 'X' could be made const reference"
                        elif error == "1762":
                            error += ": Member function could be made const"
                        elif error == "1774":
                            error += ": Could use dynamic_cast to downcast polymorphic type"
                        elif error == "1914":
                            error += ": Default constructor 'X' not referenced"
                        elif error == "1916":
                            error += ": Ellipsis encountered"
                        elif error == "1923":
                            error += ": Macro 'X' could become const variable"
                        elif error == "1924":
                            error += ": C-style cast"
                        elif error == "1960":
                            error += ": Violates MISRA C++ 2008 Required Rule"

                        errors_dict[error + "|" + comment.strip(' \n\r')].add(src_file + ":" + src_line)
                    else:
                        print("Could not parse suppression: " + suppression)
            elif "//lint +e" not in code_line and "//lint -sem" not in code_line:
                print("Could not parse code: " + code_line)

    if print_output:
        for suppression,occurences in sorted(errors_dict.items()):
            parts = suppression.split("|");
            error = parts[0]
            comment = parts[1]

            print(error)
            for occurence in sorted(occurences):
                print("\t" + occurence)
            print("\t\t" + comment)

    print("Writing: " + dst_file)

    if dst_file.endswith(".csv"):
        write_csv_file(errors_dict, dst_file)
    elif dst_file.endswith(".xlsx"):
        write_excel_file(errors_dict, dst_file)
    else:
        print("Unsupported file suffix: " + dst_file)
        sys.exit(2)

if __name__ == "__main__":
    main()
