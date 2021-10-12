#*******************************************************************************
#
# (C) COPYRIGHT Bombardier Transportation Sweden AB, 2020
#
# We reserve all rights in this file and in the information
# contained therein. Reproduction, use or disclosure to third
# parties without written authority is strictly forbidden.
#
# DESCRIPTION:
# Script that collects author and reviewers from git in order to check the
# independency between them.
#
#******************************************************************************

#******************************************************************************
#
# REVISION HISTORY :
#
# Date          Name        Changes
# ---------------------------------------------------------------------------
# 2020-09-22    csundin     Created
#
#******************************************************************************

#******************************************************************************
# INCLUDE FILES
#******************************************************************************
import os
import sys
import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment


def atoi(string, index):
    number = 0

    while index < len(string):
        char = ord(string[index])
        if char >= 0x30 and char <= 0x39:
            number = (char - 0x30) + number * 10
        else:
            break
        index += 1

    return number


def getChangeNumber(url):
    index = len(url)

    while index > 0:
        char = ord(url[index - 1])
        if char < 0x30 or char > 0x39:
            break
        index -= 1

    return atoi(url, index)


def writeExcelFile(authorDict, reviewDict, commentDict, output_filename):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50

    alignment=Alignment(horizontal='general',
                        vertical='top',
                        text_rotation=0,
                        wrap_text=True,
                        shrink_to_fit=False,
                        indent=0)

    ws.append(['Change', 'Author(s)', 'Reviewer(s)', 'Comment'])
    line_no = 1

    for change,authors in sorted(authorDict.items(), key=lambda t: getChangeNumber(t[0])):
        authorList = ""
        reviewList = ""
        comment = ""

        for author in sorted(authors):
            if authorList is not "":
                authorList += "\n"
            authorList += author

        if change in reviewDict:
            reviewers = reviewDict[change]
            for reviewer in sorted(reviewers):
                if reviewList is not "":
                    reviewList += "\n"
                reviewList += reviewer

                if reviewer in authors:
                    comment = "ERROR: Author-reviewer conflict"
                    print(comment + " for change " + change)

        if reviewList is "":
            comment = "ERROR: No independent reviewer found"
            print(comment + " for change " + change)

        if change in commentDict:
            if comment is not "":
                comment += "\n"
            comment = commentDict[change]

        ws.append([change, authorList, reviewList, comment])
        line_no += 1

        ws['A'+str(line_no)].alignment = alignment;
        ws['B'+str(line_no)].alignment = alignment;
        ws['C'+str(line_no)].alignment = alignment;
        ws['D'+str(line_no)].alignment = alignment;

    wb.save(output_filename)


def findAuthorsAndReviewers(sourcePath, startDate, userAtHost, outputName, verbose):
    authorDict = defaultdict(set)
    reviewDict = defaultdict(set)
    commentDict = {}
    removeDuplicateReviews = True
    reviewedAgain = "Reviewed again to ensure independency"

    command = "git log --since " + startDate + " " + sourcePath + " | grep \"^commit\" | sed -e s/commit//"
    if verbose:
        print("\nRunning:  " + command)

    commitsFile = os.popen(command)
    while True:
        commitLine = commitsFile.readline()
        if not commitLine: break
        commitLine = commitLine.rstrip()

        if verbose:
            print("\nCommit:  " + commitLine)

        # See https://gerrit-review.googlesource.com/Documentation/cmd-query.html
        command = "ssh -p 29418 " + userAtHost + " gerrit query --all-approvals --comments --format JSON " + commitLine
        if verbose:
            print("Running:  " + command)

        approvalsRead = False
        approvalsFile = os.popen(command)
        while True:
            approvalLine = approvalsFile.readline()
            if not approvalLine: break
            approvalData = json.loads(approvalLine.rstrip())

            if "project" in approvalData and approvalData["project"] != "aos_bhp" and approvalData["status"] == "MERGED":
                approvalsRead = True
                isAnyPatchSubmitted = False

                patchSets = approvalData["patchSets"]
                for patchSet in patchSets:
                    if "approvals" in patchSet:
                        approvals = patchSet["approvals"]
                        for approval in approvals:
                            if approval["type"] == "SUBM":
                                isAnyPatchSubmitted = True

                change = approvalData["url"]
                owner = approvalData["owner"]["name"]

                if isAnyPatchSubmitted:
                    authorDict[change].add(owner)

                    if verbose:
                        print("Change:   " + change)
                        print("Owner:    " + owner)

                    for patchSet in patchSets:
                        if patchSet["kind"] != "TRIVIAL_REBASE" and patchSet["kind"] != "NO_CHANGE" and patchSet["kind"] != "NO_CODE_CHANGE":
                            author = patchSet["author"]["name"]
                            uploader = patchSet["uploader"]["name"]

                            authorDict[change].add(uploader)

                            if author != "BHP User":
                                authorDict[change].add(author)

                            if verbose:
                                print("Author:   " + author)
                                print("Uploader: " + uploader)

                    for patchSet in patchSets:
                        if "approvals" in patchSet:
                            isSubmitted = False
                            reviewers = set()

                            approvals = patchSet["approvals"]
                            for approval in approvals:
                                if approval["type"] == "SUBM":
                                    isSubmitted = True

                                elif approval["type"] == "Code-Review" and approval["value"] == "2":
                                    reviewer = ""
                                    approvedBy = approval["by"]

                                    if "name" in approvedBy:
                                        reviewer = approvedBy["name"]

                                    if reviewer != "":
                                        reviewers.add(reviewer)
                                    else:
                                        print("WARNING: Unknown reviewer for change " + change)

                                    if verbose:
                                        print("Reviewer: " + reviewer)

                            if isSubmitted:
                                # This is enormously stupid: patch set comments are NOT stored under each
                                # patch set, so we have to find the right comment by searching for a string
                                patchSetName = "Patch Set " + patchSet["number"]

                                comments = approvalData["comments"]
                                for comment in comments:
                                    commentMessage = comment["message"]

                                    if patchSetName in commentMessage and reviewedAgain in commentMessage:
                                        reviewer = comment["reviewer"]["name"]
                                        reviewers.add(reviewer)
                                        commentDict[change] = reviewedAgain

                                        if verbose:
                                            print("Reviewer: " + reviewer)

                            for reviewer in reviewers:
                                if removeDuplicateReviews and isSubmitted and reviewer in authorDict[change] and len(reviewers) > 1:
                                    print("INFO: Skipped duplicate reviewer " + reviewer + " for change " + change)
                                else:
                                    reviewDict[change].add(reviewer)

                else:
                    print("INFO: Change " + change + " was reverted")

            elif "type" in approvalData:
                approvalsRead = True

        if not approvalsRead:
            print("Cannot read response from Gerrit")
            sys.exit(1)

    writeExcelFile(authorDict, reviewDict, commentDict, outputName)


def main():
    index = 1
    verbose = False

    if index < len(sys.argv) and sys.argv[index] == "--verbose":
        verbose = True
        index += 1

    if len(sys.argv) != index + 4:
        print("Usage: python checkIndependency.py [--verbose] <path to source code> <starting date (YYYY-MM-DD)> <user name@gerrit host> <output file>")
        sys.exit(1)

    sourcePath = sys.argv[index]
    startDate  = sys.argv[index + 1]
    userAtHost = sys.argv[index + 2]
    outputName = sys.argv[index + 3]

    findAuthorsAndReviewers(sourcePath, startDate, userAtHost, outputName, verbose)


if __name__ == "__main__":
    main()
